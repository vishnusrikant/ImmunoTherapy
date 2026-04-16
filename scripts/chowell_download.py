#!/usr/bin/env python3
"""Download Chowell 2021 (Nat Biotech) supplementary data — pan-cancer ICI cohort
with NLR, albumin, platelets, HGB, BMI, TMB + response/survival outcomes.

Paper: Chowell et al. "Improved prediction of immune checkpoint blockade efficacy
across multiple cancer types." Nature Biotechnology 2021. DOI: 10.1038/s41587-021-01070-8

Two cohorts in the xlsx:
  - Training (1,184 patients, multi-institution)
  - Test     (1,714 patients, MSK — overlaps with Valero et al. 2021 Nat Comm)

Output:
  datasets/chowell_2021/chowell_training.csv
  datasets/chowell_2021/chowell_test_msk.csv
  datasets/chowell_2021/chowell_all.csv  (concatenated, with cohort flag)
  datasets/chowell_2021/supplementary_source.xlsx

NOTE: This cohort has RESPONSE + SURVIVAL outcomes, NOT irAE severity labels.
Use for (1) cross-study validation of NLR/albumin/age/cancer feature effects,
(2) enrichment of cBioPortal features, NOT for training the severity classifier.
"""

import csv
import os
import urllib.request

URL = (
    'https://static-content.springer.com/esm/'
    'art%3A10.1038%2Fs41587-021-01070-8/MediaObjects/'
    '41587_2021_1070_MOESM3_ESM.xlsx'
)

OUT_DIR = '/home/vishnusrikant/ImmunoTherapy/datasets/chowell_2021'
XLSX_PATH = os.path.join(OUT_DIR, 'supplementary_source.xlsx')


def download_xlsx():
    os.makedirs(OUT_DIR, exist_ok=True)
    if os.path.exists(XLSX_PATH):
        print(f'[skip] already downloaded: {XLSX_PATH}')
        return
    print(f'[download] {URL}')
    req = urllib.request.Request(URL, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=120) as r, open(XLSX_PATH, 'wb') as f:
        f.write(r.read())
    print(f'[saved] {XLSX_PATH} ({os.path.getsize(XLSX_PATH):,} bytes)')


def sheet_to_csv(sheet, out_path, cohort_tag):
    """Export an openpyxl worksheet to CSV, dropping trailing blank columns
    and rows, and adding a `cohort` tag column."""
    # collect header row and find last meaningful column
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0
    header = list(rows[0])
    last_col = max(i for i, v in enumerate(header) if v not in (None, ''))
    header = header[: last_col + 1]
    # sanitize header names (keep readable, drop encoded descriptors)
    header_clean = []
    for h in header:
        s = str(h).strip()
        # strip "(1:Yes; 0:No)" style annotations
        if ' (' in s:
            s = s.split(' (', 1)[0]
        header_clean.append(s)

    written = 0
    with open(out_path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(header_clean + ['cohort'])
        for row in rows[1:]:
            row = list(row)[: last_col + 1]
            if all(v is None or v == '' for v in row):
                continue  # skip blank
            if row[0] is None:
                continue  # no sample id
            w.writerow(row + [cohort_tag])
            written += 1
    print(f'[wrote] {out_path} — {written} rows x {len(header_clean) + 1} cols')
    return written


def build_combined(train_path, test_path, out_path):
    """Concatenate training + test into a single long file with matching columns."""
    with open(train_path) as f:
        train = list(csv.DictReader(f))
    with open(test_path) as f:
        test = list(csv.DictReader(f))
    # Union of keys (test has 2 extra columns: Albumin_Chemo, ??? — check and keep all)
    keys = []
    seen = set()
    for row in train + test:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    with open(out_path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for row in train + test:
            w.writerow({k: row.get(k, '') for k in keys})
    print(
        f'[wrote] {out_path} — {len(train) + len(test)} rows '
        f'({len(train)} training + {len(test)} test) x {len(keys)} cols'
    )


def main():
    download_xlsx()

    import openpyxl  # imported lazily so the script fails cleanly if missing

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    assert 'Training' in wb.sheetnames and 'Test' in wb.sheetnames, \
        f'unexpected sheet names: {wb.sheetnames}'

    train_path = os.path.join(OUT_DIR, 'chowell_training.csv')
    test_path = os.path.join(OUT_DIR, 'chowell_test_msk.csv')
    all_path = os.path.join(OUT_DIR, 'chowell_all.csv')

    sheet_to_csv(wb['Training'], train_path, cohort_tag='training')
    sheet_to_csv(wb['Test'], test_path, cohort_tag='test_msk')
    build_combined(train_path, test_path, all_path)

    print('\n[done] Chowell 2021 cohort downloaded.')


if __name__ == '__main__':
    main()
